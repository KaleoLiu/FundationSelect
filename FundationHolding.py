import os
import re
from datetime import date, datetime, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed

import openpyxl
import requests
from bs4 import BeautifulSoup
import unicodedata

from FoundationSelect import FoundSelect


SITCA_URL = "https://www.sitca.org.tw/ROC/Industry/IN2629.aspx?pid=IN22601_04"
DEFAULT_CLASS_VALUE = "AA1"  # 類型預設 AA1（即使下拉 disabled）

# 一次處理幾檔基金（由 FundRanking_* 中前幾名決定）
MAX_FUNDS_PER_RUN = 10

# 同一時間最多對 SITCA 發出幾個請求
MAX_CONCURRENT_REQUESTS = 6


def _clean(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "").strip())


def normalize_fund_name(name: str) -> str:
    """
    將基金名稱做寬鬆正規化：
    - 去除前後空白、連續空白
    - 全形/半形統一（例如 'ｅ' -> 'e'）
    - 移除所有空白
    """
    s = _clean(name)
    s = unicodedata.normalize("NFKC", s)
    s = s.replace(" ", "")
    return s


# SITCA 常見基金名稱後綴（括號說明），剝除後可與 Excel 排名名稱對齊
FUND_NAME_SUFFIX_PATTERN = re.compile(
    r"\s*[(\（].*?(?:本基金之?配息來源可能為本金|配息來源可能為本金).*?[)\）]?\s*$"
)

# Excel 排名常見「級別」後綴：A累積型(台幣)、B月配型(台幣)、-A不配息、-I不配息 等，持股與主基金相同
SHARE_CLASS_SUFFIX_PATTERN = re.compile(
    r"(?:[ABNI]?(?:累積型|月配型|不配息)|-[AI]不配息)(?:[(\（]台幣[)\）])?\s*$",
    re.IGNORECASE,
)
CURRENCY_SUFFIX_PATTERN = re.compile(r"[(\（]台幣[)\）]\s*$")


def fund_name_base(normalized: str) -> str:
    """剝除 SITCA 常見括號後綴，得到基底名稱（如「台中銀數位時代基金(本基金…)」→「台中銀數位時代基金」）"""
    return FUND_NAME_SUFFIX_PATTERN.sub("", normalized).strip()


def excel_fund_name_to_lookup_base(fund_name: str) -> str:
    """
    從 Excel 排名基金名稱推導「主基金名」用於 SITCA 對照。
    同一主基金之不同級別（A累積型、B月配型、-A不配息、-I不配息）持股相同，對到同一筆 SITCA 即可。
    例：「台中銀數位時代基金B月配型(台幣)(本基金之配息…」→「台中銀數位時代基金」
        「台新2000高科技基金-A不配息」→「台新2000高科技基金」
    """
    s = normalize_fund_name(fund_name)
    s = FUND_NAME_SUFFIX_PATTERN.sub("", s).strip()
    s = CURRENCY_SUFFIX_PATTERN.sub("", s).strip()
    s = SHARE_CLASS_SUFFIX_PATTERN.sub("", s).strip()
    return s


def format_month_to_yyyymm(month_text: str) -> str:
    """
    將 '2025 年 12 月' 轉成 '202512'
    """
    m = re.search(r"(\d{4})\s*年\s*(\d{2})\s*月", month_text)
    if not m:
        raise ValueError(f"Unsupported month format: {month_text}")
    return f"{m.group(1)}{m.group(2)}"


def parse_company_code_from_excel_text(company_text: str) -> str:
    """
    '元大投信' -> 嘗試在 SITCA 公司下拉中找對應 value（例如 A0005）
    """
    return _clean(company_text)


def get_hidden_inputs(soup: BeautifulSoup) -> dict[str, str]:
    hidden_names = ["__VIEWSTATE", "__VIEWSTATEGENERATOR", "__EVENTVALIDATION"]
    data: dict[str, str] = {}
    for name in hidden_names:
        el = soup.find("input", {"name": name})
        if el and el.get("value") is not None:
            data[name] = el["value"]
    missing = [n for n in hidden_names if n not in data]
    if missing:
        raise RuntimeError(f"缺少必要 hidden 欄位: {missing}")
    return data


class SitcaFastClient:
    def __init__(self):
        self.session = requests.Session()
        self.session.headers.update(
            {
                "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120 Safari/537.36",
                "Accept-Language": "zh-TW,zh;q=0.9,en;q=0.8",
            }
        )
        self._company_map: dict[str, str] | None = None  # '元大投信' -> 'A0005'

    def close(self) -> None:
        """釋放 Session 連線池，避免爬蟲失敗時連線殘留"""
        self.session.close()

    def _get_with_retry(self, url: str, retries: int = 3, timeout: int = 30):
        last_exc: Exception | None = None
        for _ in range(retries):
            try:
                resp = self.session.get(url, timeout=timeout)
                resp.raise_for_status()
                return resp
            except Exception as e:
                last_exc = e
        assert last_exc is not None
        raise last_exc

    def _post_with_retry(
        self, url: str, data: dict[str, str], retries: int = 3, timeout: int = 30
    ):
        last_exc: Exception | None = None
        for _ in range(retries):
            try:
                resp = self.session.post(url, data=data, timeout=timeout)
                resp.raise_for_status()
                return resp
            except Exception as e:
                last_exc = e
        assert last_exc is not None
        raise last_exc

    def _get_company_map(self) -> dict[str, str]:
        if self._company_map is not None:
            return self._company_map

        r = self._get_with_retry(SITCA_URL, retries=3, timeout=30)
        soup = BeautifulSoup(r.text, "html.parser")

        select = soup.find("select", {"id": "ctl00_ContentPlaceHolder1_ddlQ_Comid"})
        if not select:
            raise RuntimeError(
                "找不到公司下拉選單 ctl00_ContentPlaceHolder1_ddlQ_Comid"
            )

        mapping: dict[str, str] = {}
        for opt in select.find_all("option"):
            value = (opt.get("value") or "").strip()
            text = _clean(opt.get_text())
            # text 可能是 "A0005 元大投信"，取最後的公司名
            name = _clean(re.sub(r"^A\d{4}\s*", "", text))
            if value and name:
                mapping[name] = value

        if not mapping:
            raise RuntimeError("公司下拉選單解析失敗（mapping 為空）")

        self._company_map = mapping
        return mapping

    def company_name_to_value(self, company_name: str) -> str:
        company_name = parse_company_code_from_excel_text(company_name)
        m = self._get_company_map()
        if company_name in m:
            return m[company_name]
        # 寬鬆匹配：用包含判斷
        for k, v in m.items():
            if company_name in k or k in company_name:
                return v
        raise KeyError(f"找不到投信公司對應代碼: {company_name}")

    def fetch_company_month_table(
        self, yyyymm: str, company_value: str
    ) -> dict[str, list[list[str]]]:
        """
        取得某個月份 + 某家公司（A0005）下的整張「月前十大」資料，
        回傳：fund_name -> rows
        rows 每列為：
            [名次, 標的種類, 標的代號, 標的名稱, 金額, 擔保機構, 次順位債券, 受益權單位數, 占比]
        """
        # 1) GET 取得 hidden
        r = self._get_with_retry(SITCA_URL, retries=3, timeout=30)
        soup = BeautifulSoup(r.text, "html.parser")
        hidden = get_hidden_inputs(soup)

        # 2a) 先 postback 年月
        payload_ym: dict[str, str] = {
            **hidden,
            "__EVENTTARGET": "ctl00$ContentPlaceHolder1$ddlQ_YM",
            "__EVENTARGUMENT": "",
            "ctl00$ContentPlaceHolder1$ddlQ_YM": yyyymm,
            "ctl00$ContentPlaceHolder1$ddlQ_Class": DEFAULT_CLASS_VALUE,
        }
        r_ym = self._post_with_retry(SITCA_URL, data=payload_ym, retries=3, timeout=30)
        soup_ym = BeautifulSoup(r_ym.text, "html.parser")
        hidden2 = get_hidden_inputs(soup_ym)

        # 2b) 查詢（公司）
        payload_query: dict[str, str] = {
            **hidden2,
            "__EVENTTARGET": "",
            "__EVENTARGUMENT": "",
            "ctl00$ContentPlaceHolder1$ddlQ_YM": yyyymm,
            "ctl00$ContentPlaceHolder1$rdo1": "rbComid",
            "ctl00$ContentPlaceHolder1$ddlQ_Comid": company_value,
            "ctl00$ContentPlaceHolder1$ddlQ_Class": DEFAULT_CLASS_VALUE,
            "ctl00$ContentPlaceHolder1$BtnQuery": "查詢",
        }
        r2 = self._post_with_retry(SITCA_URL, data=payload_query, retries=3, timeout=30)
        soup2 = BeautifulSoup(r2.text, "html.parser")

        # 3) 找結果 table（用純文字判斷避免 HTML 不嚴謹）
        result_table = None
        for table in soup2.find_all("table"):
            text = table.get_text(" ", strip=True)
            if ("基金名稱" in text) and ("標的種類" in text) and ("名次" in text):
                result_table = table
                break
        if not result_table:
            raise RuntimeError(
                "找不到結果表格（未找到含 基金名稱/標的種類/名次 的 table）"
            )

        fund_rows: dict[str, list[list[str]]] = {}
        current_fund: str | None = None

        for tr in result_table.find_all("tr")[1:]:
            tds = tr.find_all("td")
            if not tds:
                continue

            # row 可能含基金名稱（rowspan=10）
            if len(tds) >= 10:
                current_fund = _clean(tds[0].get_text())
                cells = tds[1:]
            else:
                cells = tds

            if not current_fund:
                continue

            # cells 預期至少 9 欄：名次..占比
            if len(cells) < 9:
                continue

            row = [_clean(c.get_text()) for c in cells[:9]]
            fund_rows.setdefault(current_fund, []).append(row)

        return fund_rows


# 最近一個月份：數據寫入Excel
def truncate_sheet_name(sheet_name, max_length=31):
    if len(sheet_name) <= max_length:
        return sheet_name
    return sheet_name[:max_length]


def write_recrrent_month_excel(data, sheet_name, header, filename):
    workbook = openpyxl.load_workbook(filename)
    sheet_name = truncate_sheet_name(sheet_name)
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    worksheet = workbook.create_sheet(title=sheet_name)

    # Insert the description at the top of the first row
    description = ["最近一個月資料"]
    worksheet.append(description)

    # 首先添加標頭
    worksheet.append(header)
    for row in data:
        worksheet.append(row)
    workbook.save(filename)


# 最近兩個月份：數據寫入Excel
def write_previous_month_excel(data, sheet_name, header, filename):
    workbook = openpyxl.load_workbook(filename)
    sheet_name = truncate_sheet_name(sheet_name)
    worksheet = workbook[sheet_name]

    # Add description to the first cell in the first row
    worksheet.cell(row=1, column=10, value="最近兩個月資料")

    # 添加標頭至第十行開頭
    for col_idx, header_title in enumerate(header, start=10):
        worksheet.cell(row=2, column=col_idx, value=header_title)

    # 從第一行第10列開始寫入多行數據
    for row_idx, row_data in enumerate(data, start=3):
        for col_idx, value in enumerate(row_data, start=10):
            worksheet.cell(row=row_idx, column=col_idx, value=value)

    workbook.save(filename)


# 兩個月份比較：數據寫入Excel
def write_compare_fundation_excel(increase, new, delete, sheet_name, filename):
    workbook = openpyxl.load_workbook(filename)
    sheet_name = truncate_sheet_name(sheet_name)
    worksheet = workbook[sheet_name]

    # 從第15行第1列開始寫入多行數據
    for row_idx, row_data in enumerate(increase, start=15):
        for col_idx, value in enumerate(row_data, start=1):
            worksheet.cell(row=row_idx, column=col_idx, value=value)

    # 從第15行第5列開始寫入多行數據
    for row_idx, row_data in enumerate(new, start=15):
        for col_idx, value in enumerate(row_data, start=5):
            worksheet.cell(row=row_idx, column=col_idx, value=value)

    # 從第15行第10列開始寫入多行數據
    for row_idx, row_data in enumerate(delete, start=15):
        for col_idx, value in enumerate(row_data, start=9):
            worksheet.cell(row=row_idx, column=col_idx, value=value)
    workbook.save(filename)


def compare_fundation_in_different_months(recent_months, previous_months):
    # recent/previous: list rows; each row [名次, 標的種類, 標的代號, 標的名稱, 金額, ...]
    re_month_arrey = [[row[2], row[3], row[4]] for row in recent_months]
    pr_month_arrey = [[row[2], row[3], row[4]] for row in previous_months]

    # 增加持股
    increase_fundations = ["增加持股", ["標的代號", "標的名稱", "金額", "差額"]]
    re_map = {r[0]: r for r in re_month_arrey if r and r[0]}
    for pr in pr_month_arrey:
        if not pr or not pr[0]:
            continue
        if pr[0] in re_map:
            re_row = re_map[pr[0]]
            try:
                pr_val = float((pr[2] or "0").replace(",", ""))
                re_val = float((re_row[2] or "0").replace(",", ""))
            except Exception:
                continue
            if pr_val <= re_val:
                increase_fundations.append([pr[0], pr[1], pr[2], re_val - pr_val])

    # 新增持股
    new_fundations = ["新增持股", ["標的代號", "標的名稱", "金額"]]
    pr_codes = {r[0] for r in pr_month_arrey if r and r[0]}
    for re_row in re_month_arrey:
        if re_row and re_row[0] and re_row[0] not in pr_codes:
            new_fundations.append(re_row)

    # 剔除持股
    delete_fundations = ["剔除持股", ["標的代號", "標的名稱", "金額"]]
    re_codes = {r[0] for r in re_month_arrey if r and r[0]}
    for pr in pr_month_arrey:
        if pr and pr[0] and pr[0] not in re_codes:
            delete_fundations.append(pr)

    return increase_fundations, new_fundations, delete_fundations


def calculate_business_days_in_month(year, month):
    """計算一個月中的工作日數量"""
    first_day = date(year, month, 1)
    last_day = (
        date(year, month + 1, 1) - timedelta(days=1)
        if month != 12
        else date(year, 12, 31)
    )

    # 生成該月的日期範圍
    all_days = [date(year, month, day) for day in range(1, last_day.day + 1)]

    # 計算工作日（排除週六和週日）
    weekdays = [d.weekday() for d in all_days]
    business_days = [d for d, wd in zip(all_days, weekdays) if wd < 5]

    return business_days


def get_current_and_previous_months(today):
    """根據今天是當月第幾個工作日來獲取當前月份和前一個月份"""
    current_month_business_days = calculate_business_days_in_month(
        today.year, today.month
    )
    # 檢查今天是否為工作日，如果不是則找到最接近的前一個工作日
    if today not in current_month_business_days:
        # 獲取比今天早的工作日列表
        earlier_business_days = [d for d in current_month_business_days if d < today]
        # 選擇最接近今天的工作日
        closest_business_day = (
            max(earlier_business_days) if earlier_business_days else None
        )
        if closest_business_day:
            todays_index_in_business_days = (
                current_month_business_days.index(closest_business_day) + 1
            )
        else:
            raise ValueError("No earlier business days found in the current month.")
    else:
        todays_index_in_business_days = current_month_business_days.index(today) + 1
    # 計算當前月份的第一天
    first_day_of_current_month = date(today.year, today.month, 1)

    if todays_index_in_business_days > 9:
        # 計算前一個月的最後一天
        current_month = first_day_of_current_month + timedelta(days=1)
    else:
        # 如果今天不是第十個工作日之後
        current_month = first_day_of_current_month - timedelta(days=1)

    days = current_month
    # 前一個月：特別處理當前月份為1月的情況
    if days.month == 1:
        previous_1month = date(days.year - 1, 12, 1)
    else:
        previous_1month = date(days.year, days.month - 1, 1)

    days = previous_1month
    # 前兩個月：特別處理當前月份為1月的情況
    if days.month == 1:
        previous_2month = date(days.year - 1, 12, 1)
    else:
        previous_2month = date(days.year, days.month - 1, 1)

    # 格式化為“年 年 月 月”格式
    formatted_previous_1month = previous_1month.strftime("%Y 年 %m 月")
    formatted_previous_2month = previous_2month.strftime("%Y 年 %m 月")

    return formatted_previous_1month, formatted_previous_2month


def process_fundation_fast(
    fund_row,
    months_text: list[str],
    client: SitcaFastClient,
    cache: dict[tuple[str, str], dict[str, list[list[str]]]],
    filename: str,
):
    fund_name = fund_row[1]["基金名稱"]
    norm_target = normalize_fund_name(fund_name)
    fund_company_name = fund_row[1]["基金公司"]
    company_value = client.company_name_to_value(fund_company_name)

    headers = [
        "名次",
        "標的種類",
        "標的代號",
        "標的名稱",
        "金額",
        "擔保機構",
        "次順位債券",
        "受益權單位數",
        "基金淨資產價值之比例",
    ]

    results: list[list[list[str]]] = []
    for mtext in months_text:
        yyyymm = format_month_to_yyyymm(mtext)
        key = (yyyymm, company_value)
        table = cache.get(key)
        if table is None:
            table = client.fetch_company_month_table(yyyymm, company_value)
            cache[key] = table

        # 先精確，再做名稱正規化／主基金名匹配（同主基金不同級別持股相同，對到同一筆 SITCA）
        rows = table.get(fund_name, [])
        if not rows:
            # 建立對照：正規化名、基底名（剝除「(本基金配息…)」）→ SITCA 原始 key
            norm_map: dict[str, str] = {}
            for n in table.keys():
                n_norm = normalize_fund_name(n)
                norm_map[n_norm] = n
                base = fund_name_base(n_norm)
                if base and base != n_norm:
                    norm_map[base] = n
            # 依序嘗試：正規化名、Excel 主基金名（剝除 A累積型/B月配型/-A不配息 等）、前綴、包含
            lookup_keys = [
                norm_target,
                excel_fund_name_to_lookup_base(fund_name),
            ]
            for key in lookup_keys:
                if key and key in norm_map:
                    rows = table.get(norm_map[key], [])
                    if rows:
                        break
            if not rows:
                # fallback 1：SITCA 名稱以 Excel 名稱為前綴（或主基金名為前綴）
                for key in lookup_keys:
                    if not key:
                        continue
                    candidates = [
                        orig
                        for n_norm, orig in norm_map.items()
                        if n_norm.startswith(key)
                    ]
                    if candidates:
                        best = max(candidates, key=lambda o: len(normalize_fund_name(o)))
                        rows = table.get(best, [])
                        if rows:
                            break
            if not rows:
                # fallback 2：包含關係
                for key in lookup_keys:
                    if not key:
                        continue
                    for n_norm, orig in norm_map.items():
                        if key in n_norm or n_norm in key:
                            rows = table.get(orig, [])
                            if rows:
                                break
                    if rows:
                        break
        # 只留前十筆
        results.append(rows[:10])

    # 保護：確保兩個月份都有資料結構
    current_rows = results[0] if len(results) > 0 else []
    prev_rows = results[1] if len(results) > 1 else []

    write_recrrent_month_excel(current_rows, fund_name, headers, filename)
    write_previous_month_excel(prev_rows, fund_name, headers, filename)

    increase, new, delete = compare_fundation_in_different_months(
        current_rows, prev_rows
    )
    write_compare_fundation_excel(increase, new, delete, fund_name, filename)


def FundationHolding(filename):
    today = date.today()
    print("今天日期:", today)

    if not os.path.exists(filename):
        raise FileNotFoundError(
            f"文件不存在: {filename}（請先執行 FundationTaiwan 建檔）"
        )

    formatted_current_month, formatted_previous_month = get_current_and_previous_months(
        today
    )
    print("當前月份:", formatted_current_month)
    print("前一個月份:", formatted_previous_month)

    months_text = [formatted_current_month, formatted_previous_month]

    client = SitcaFastClient()
    try:
        cache: dict[tuple[str, str], dict[str, list[list[str]]]] = {}

        # 從當日排名檔中選出要處理的基金（預設取前 MAX_FUNDS_PER_RUN 檔）
        Fundations = FoundSelect(MAX_FUNDS_PER_RUN, file_path=filename)
        Fundations = Fundations.head(MAX_FUNDS_PER_RUN)

        # 先把所有需要的 (月份, 公司) key 列出來，並行預抓，後續每支基金只做 dict lookup
        needed_keys: set[tuple[str, str]] = set()
        for _, row in Fundations.iterrows():
            company_value = client.company_name_to_value(row["基金公司"])
            for mtext in months_text:
                needed_keys.add((format_month_to_yyyymm(mtext), company_value))

        print(
            f"預計抓取 {len(needed_keys)} 個 (月份, 公司) 組合，使用 requests（不開瀏覽器）"
        )

        with ThreadPoolExecutor(max_workers=MAX_CONCURRENT_REQUESTS) as ex:
            futs = {
                ex.submit(client.fetch_company_month_table, yyyymm, comp): (yyyymm, comp)
                for (yyyymm, comp) in needed_keys
            }
            for fut in as_completed(futs):
                key = futs[fut]
                try:
                    cache[key] = fut.result()
                except Exception as e:
                    print(f"抓取失敗 {key}: {e}")
                    cache[key] = {}

        # 寫入每支基金
        for fund_row in Fundations.iterrows():
            process_fundation_fast(fund_row, months_text, client, cache, filename)
    finally:
        client.close()


if __name__ == "__main__":

    # Get the current date in the format YYYY-MM-DD
    current_date = datetime.now().strftime("%Y-%m-%d")
    # Construct the filename with the current date appended
    filename = f"FundRanking_{current_date}.xlsx"

    FundationHolding(filename)
