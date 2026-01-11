from datetime import datetime
from openpyxl import load_workbook
from FundationTaiwan import FundationTaiwan
from FundationHolding import FundationHolding
from FundationCompare import FundationCompare

if __name__ == "__main__":

    # Get the current date in the format YYYY-MM-DD
    current_date = datetime.now().strftime("%Y-%m-%d")
    # Construct the filename with the current date appended
    filename = f"FundRanking_{current_date}.xlsx"

    FundationTaiwan(filename)
    FundationHolding(filename)
    FundationCompare(filename)

    # Load the workbook
    wb = load_workbook(filename)
    # 删除默认创建的空白工作表
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Save the workbook
    wb.save(filename)
