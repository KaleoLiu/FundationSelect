import os
import ssl
import threading
import time
from datetime import datetime

import openpyxl
import pandas as pd
import requests
from bs4 import BeautifulSoup
from requests.adapters import HTTPAdapter

try:
    import certifi
except ImportError:
    certifi = None


# Fubon「基金類型」頁面基底 URL
BASE_URL = "https://fubon-ebrokerdj.fbs.com.tw"
FUND_TYPE_PATHS = {
    # 這些路徑是從原本 Selenium 選單觀察得來，改成直接走 requests 加速
    "國內股票開放型科技類": "/w/wq/wq02_ET001001_801.djhtm",
    "國內股票開放型中小型": "/w/wq/wq02_ET001004_801.djhtm",
    "國內股票開放型一般股票型": "/w/wq/wq02_ET001005_801.djhtm",
}


def _create_secure_session() -> requests.Session:
    """建立帶安全 SSL context 的 requests Session（避免 Python 3.13 嚴格 SKI 問題）"""

    def create_secure_ssl_context():
        ctx = ssl.SSLContext(ssl.PROTOCOL_TLS_CLIENT)
        if certifi:
            ctx.load_verify_locations(certifi.where())
        ctx.verify_mode = ssl.CERT_REQUIRED
        ctx.check_hostname = True
        return ctx

    class SecureHTTPAdapter(HTTPAdapter):
        def init_poolmanager(self, *args, **kwargs):
            kwargs["ssl_context"] = create_secure_ssl_context()
            return super().init_poolmanager(*args, **kwargs)

    session = requests.Session()
    session.headers.update(
        {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120 Safari/537.36"
        }
    )
    secure_adapter = SecureHTTPAdapter()
    session.mount("https://", secure_adapter)
    return session


def get_data(fund_type: str, retries: int = 3) -> str:
    """使用 requests 直接抓取指定基金類型的績效排行頁面 HTML；503/502/504 時自動重試"""
    if fund_type not in FUND_TYPE_PATHS:
        raise ValueError(f"未知的 fund_type: {fund_type}")

    fund_type_url = BASE_URL + FUND_TYPE_PATHS[fund_type]
    last_exc = None
    for attempt in range(retries):
        session = _create_secure_session()
        try:
            resp = session.get(fund_type_url, timeout=30)
            if resp.status_code in (502, 503, 504) and attempt < retries - 1:
                time.sleep(2 + attempt * 2)  # 2s, 4s, 6s
                continue
            resp.raise_for_status()
            return resp.text
        except Exception as e:
            last_exc = e
            if attempt < retries - 1:
                time.sleep(2 + attempt * 2)
        finally:
            session.close()
    if last_exc:
        raise last_exc
    raise RuntimeError(f"抓取失敗: {fund_type}")


def parse_table(html):
    # 使用 BeautifulSoup 解析網頁內容
    soup = BeautifulSoup(html, "html.parser")
    table = soup.find("table")
    rows = table.find_all("tr")

    # data = pd.DataFrame()

    all_row_data = []  # 初始化用于存储所有行数据的列表

    # 解析表格的每一行，將資料保存到 DataFrame
    for row in rows:
        cols = row.find_all("td")
        if cols:
            row_data = [col.text.strip() for col in cols]
            # row_df = pd.DataFrame([row_data])
            # data = pd.concat([data, row_df], ignore_index=True)
            all_row_data.append(row_data)

    # 直接从 all_row_data 列表创建 DataFrame
    data = pd.DataFrame(all_row_data)

    # 定位表格中“上述資料”的行
    start_row = data[data.iloc[:, 0].str.contains("上述資料")].index[0]
    # 保留表格的前幾行，並且只保留前10列
    data = data.iloc[:start_row]
    data = data.iloc[2:, :10]

    return data


def create_empty_excel(file_name):
    # 創建一個空的 Excel 文件
    workbook = openpyxl.Workbook()
    workbook.save(file_name)


write_lock = threading.Lock()


def write_to_excel(data, sheet_name, File_name):
    # 获取锁
    with write_lock:
        workbook = openpyxl.load_workbook(File_name)
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]
        worksheet = workbook.create_sheet(title=sheet_name)

        for row in data.values:
            worksheet.append(row.tolist())

        workbook.save(File_name)


# FundationGet函數，這裡我們將其設計為一個可以接受參數的函數，方便多線程處理
def FundationGet(fund_type, File_name):
    # 獲取網頁內容
    html = get_data(fund_type)
    # 解析表格並獲取數據
    data = parse_table(html)
    # 將數據寫入Excel
    write_to_excel(data, fund_type, File_name)


def FundationTaiwan(filename):

    # 定義需要處理的基金類型
    fund_types = [
        "國內股票開放型科技類",
        "國內股票開放型一般股票型",
        "國內股票開放型中小型",
    ]

    # If a file with the same name already exists, remove it
    if os.path.exists(filename):
        os.remove(filename)

    # Create an empty Excel file with the new filename
    create_empty_excel(filename)

    """# 如果已存在該Excel，則先移除
    if os.path.exists("FundRanking.xlsx"):
        os.remove("FundRanking.xlsx")
    # 創建一個空的Excel文件
    create_empty_excel("FundRanking.xlsx")
    """
    # 定義一個儲存多線程的list
    threads = []

    # 為每種基金類型開啟一個新的線程
    for fund_type in fund_types:
        # 創建新線程，目標函數為FundationGet，參數為fund_type
        thread = threading.Thread(target=FundationGet, args=(fund_type, filename))
        # 將新創建的線程添加到線程列表中
        threads.append(thread)
        # 開始執行新創建的線程
        thread.start()

    # 等待所有線程結束
    for thread in threads:
        thread.join()


if __name__ == "__main__":

    # Get the current date in the format YYYY-MM-DD
    current_date = datetime.now().strftime("%Y-%m-%d")

    # Construct the filename with the current date appended
    filename = f"FundRanking_{current_date}.xlsx"

    FundationTaiwan(filename)
