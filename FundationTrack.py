import time
from datetime import datetime
from io import StringIO
import requests
import pandas as pd
from openpyxl import load_workbook


def get_Fundation_Number(filename, sheet_name="Summary"):
    Fundation_codes = []
    # Load the workbook
    wb = load_workbook(filename)

    # Check if the sheet_name exists in the workbook
    if sheet_name in wb.sheetnames:
        sheet_data = wb[sheet_name]

        # Do something with sheet_data, like printing all values in the first column
        for row in sheet_data.iter_rows(min_row=2, max_col=1, values_only=True):
            Fundation_codes.append(row[0])
            # print(row[0])
        for Fundation_code in Fundation_codes:
            print(Fundation_code)
        return Fundation_codes

    else:
        print("Could not find the sheet")


def monthly_report(year, month):

    # 假如是西元，轉成民國
    if year > 1990:
        year -= 1911

    url = (
        "https://mops.twse.com.tw/nas/t21/sii/t21sc03_"
        + str(year)
        + "_"
        + str(month)
        + "_0.html"
    )
    if year <= 98:
        url = (
            "https://mops.twse.com.tw/nas/t21/sii/t21sc03_"
            + str(year)
            + "_"
            + str(month)
            + ".html"
        )

    # 偽瀏覽器
    headers = {
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.95 Safari/537.36"
    }

    # 下載該年月的網站，並用pandas轉換成 dataframe
    r = requests.get(url, headers=headers)
    r.encoding = "big5"

    dfs = pd.read_html(StringIO(r.text), encoding="big5")

    df = pd.concat([df for df in dfs if df.shape[1] <= 11 and df.shape[1] > 5])

    if "levels" in dir(df.columns):
        df.columns = df.columns.get_level_values(1)
    else:
        df = df[list(range(0, 10))]
        column_index = df.index[(df[0] == "公司代號")][0]
        df.columns = df.iloc[column_index]
    df["當月營收"] = pd.to_numeric(df["當月營收"], "coerce")
    df = df[~df["當月營收"].isnull()]
    df = df[df["公司 代號"] != "合計"]

    # 偽停頓
    time.sleep(5)

    return df


if __name__ == "__main__":

    # Get the current date in the format YYYY-MM-DD
    current_date = datetime.now().strftime("%Y-%m-%d")
    # Construct the filename with the current date appended
    filename = f"FundRanking_{current_date}.xlsx"
    sheet_name = "Summary"
    Fundation_codes = get_Fundation_Number(filename, sheet_name)
    current_year, current_month, current_day = (
        int(datetime.now().strftime("%Y")),
        int(datetime.now().strftime("%m")),
        int(datetime.now().strftime("%d")),
    )
    if current_month == 1:
        last_month = 12
        last_year = current_year - 1
    else:
        last_month = current_month - 1
        last_year = current_year
    Fundation_Revenue = monthly_report(last_year, last_month)
